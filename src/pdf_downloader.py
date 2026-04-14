"""Download ReScience C and original-manuscript PDFs and record filenames in the bibtex table.

- ReScience C papers: fetched from the direct URL in the `url` column and saved to
  ``data/pdf_cache/``.
- Original manuscripts: located by DOI via Unpaywall (open-access) with a Crossref
  fallback, and saved to ``data/pdf_original_cache/``.

After downloading, the xlsx table is updated with two new ``pdf_file`` columns — one
appended to the end of each section — that record the filename that actually made it
to disk.
"""

from __future__ import annotations

import re
import time
import warnings
from pathlib import Path

import openpyxl
import pandas as pd
import requests

warnings.filterwarnings("ignore", message="Unverified HTTPS request")

SCRIPT_DIR = Path(__file__).parent
XLSX_PATH = SCRIPT_DIR / "data" / "rescience_bibtex_table.xlsx"
RESCIENCE_CACHE = SCRIPT_DIR / "data" / "pdf_cache"
ORIGINAL_CACHE = SCRIPT_DIR / "data" / "pdf_original_cache"

# Unpaywall requires an email for API access; harmless to expose.
UNPAYWALL_EMAIL = "reproducibility@21st-publications.local"

REQUEST_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (compatible; ReproducibilityBot/1.0; "
        f"mailto:{UNPAYWALL_EMAIL})"
    ),
    "Accept": "application/pdf,*/*;q=0.8",
}

# Column layout of the xlsx (after skipping the sub-header row):
# 0:              idx
# 1..17:          RESCIENCE C section (ends at rescience_filename)
# 18..26:         ORIGINAL section    (ends at original_filename)
RESCIENCE_COLUMNS = [
    "idx",
    "rescience_authors", "rescience_title", "rescience_journal",
    "rescience_year", "rescience_month", "rescience_volume",
    "rescience_number", "rescience_doi", "rescience_url",
    "code_url", "code_doi", "data_url", "data_doi",
    "language", "domain", "keywords", "rescience_filename",
    "original_authors", "original_title", "original_journal",
    "original_year", "original_month", "original_volume",
    "original_number", "original_doi", "original_filename",
]


def load_table() -> pd.DataFrame:
    """Load the xlsx, drop the sub-header row, and apply canonical column names."""
    df = pd.read_excel(XLSX_PATH)
    df = df.iloc[1:].reset_index(drop=True)
    df.columns = RESCIENCE_COLUMNS[: len(df.columns)]
    return df


def safe_pdf_name(name: str) -> str:
    """Strip path separators and ensure a .pdf suffix."""
    name = re.sub(r"[^A-Za-z0-9._-]", "_", name.strip())
    if not name.lower().endswith(".pdf"):
        name += ".pdf"
    return name


def normalise_github_url(url: str) -> str:
    """Turn GitHub blob URLs into raw URLs so we receive the actual PDF bytes."""
    return re.sub(r"github\.com/(.+)/blob/", r"github.com/\1/raw/", url)


def download_pdf(url: str, dest: Path) -> None:
    """Download ``url`` to ``dest``. Raises if the response is not a PDF."""
    url = normalise_github_url(url)
    dest.parent.mkdir(parents=True, exist_ok=True)
    resp = requests.get(
        url,
        headers=REQUEST_HEADERS,
        timeout=120,
        verify=False,
        allow_redirects=True,
    )
    resp.raise_for_status()
    body = resp.content
    if not body.startswith(b"%PDF"):
        ctype = resp.headers.get("Content-Type", "?")
        raise ValueError(f"not a PDF (Content-Type={ctype}, {len(body)} bytes)")
    dest.write_bytes(body)


def unpaywall_pdf_url(doi: str) -> str | None:
    """Return the best open-access PDF URL for a DOI via Unpaywall, else None."""
    try:
        resp = requests.get(
            f"https://api.unpaywall.org/v2/{doi}",
            params={"email": UNPAYWALL_EMAIL},
            timeout=30,
        )
    except requests.RequestException:
        return None
    if resp.status_code != 200:
        return None
    data = resp.json()
    best = data.get("best_oa_location") or {}
    if best.get("url_for_pdf"):
        return best["url_for_pdf"]
    for loc in data.get("oa_locations") or []:
        if loc.get("url_for_pdf"):
            return loc["url_for_pdf"]
    return None


def crossref_pdf_url(doi: str) -> str | None:
    """Fallback: look for a PDF link in the Crossref record for ``doi``."""
    try:
        resp = requests.get(
            f"https://api.crossref.org/works/{doi}",
            headers={"User-Agent": REQUEST_HEADERS["User-Agent"]},
            timeout=30,
        )
    except requests.RequestException:
        return None
    if resp.status_code != 200:
        return None
    message = resp.json().get("message", {}) or {}
    for link in message.get("link") or []:
        if (link.get("content-type") or "").lower() == "application/pdf":
            return link.get("URL")
    return None


def resolve_original_pdf_url(doi: str) -> str | None:
    """Try Unpaywall first, then Crossref, to find a PDF URL for the given DOI."""
    return unpaywall_pdf_url(doi) or crossref_pdf_url(doi)


def download_rescience(df: pd.DataFrame) -> dict[int, str]:
    """Download RESCIENCE C PDFs. Returns {row_index: saved_filename}."""
    RESCIENCE_CACHE.mkdir(parents=True, exist_ok=True)
    saved: dict[int, str] = {}
    total = len(df)
    for idx, row in df.iterrows():
        url = str(row["rescience_url"]).strip()
        filename = str(row["rescience_filename"]).strip()
        if not url.startswith("http") or filename.lower() == "nan":
            print(f"[rescience {idx+1}/{total}] skip (no URL)")
            continue
        dest = RESCIENCE_CACHE / safe_pdf_name(filename)
        if dest.exists() and dest.stat().st_size > 0:
            saved[idx] = dest.name
            print(f"[rescience {idx+1}/{total}] cached -> {dest.name}")
            continue
        try:
            download_pdf(url, dest)
            saved[idx] = dest.name
            print(f"[rescience {idx+1}/{total}] saved  -> {dest.name}")
        except Exception as exc:
            print(f"[rescience {idx+1}/{total}] FAILED ({filename}): {exc}")
        time.sleep(0.4)
    return saved


def download_originals(df: pd.DataFrame) -> dict[int, str]:
    """Download original manuscripts via DOI lookups."""
    ORIGINAL_CACHE.mkdir(parents=True, exist_ok=True)
    saved: dict[int, str] = {}
    total = len(df)
    for idx, row in df.iterrows():
        doi = str(row.get("original_doi", "")).strip()
        filename = str(row.get("original_filename", "")).strip()
        if not doi or doi.lower() == "nan":
            print(f"[original {idx+1}/{total}] skip (no DOI)")
            continue
        if filename.lower() == "nan" or not filename:
            filename = f"{safe_pdf_name(doi.replace('/', '_'))}"
        dest = ORIGINAL_CACHE / safe_pdf_name(filename)
        if dest.exists() and dest.stat().st_size > 0:
            saved[idx] = dest.name
            print(f"[original {idx+1}/{total}] cached -> {dest.name}")
            continue
        pdf_url = resolve_original_pdf_url(doi)
        if not pdf_url:
            print(f"[original {idx+1}/{total}] no OA PDF for DOI {doi}")
            time.sleep(0.4)
            continue
        try:
            download_pdf(pdf_url, dest)
            saved[idx] = dest.name
            print(f"[original {idx+1}/{total}] saved  -> {dest.name}  (via {pdf_url[:60]}...)")
        except Exception as exc:
            print(f"[original {idx+1}/{total}] FAILED ({doi}): {exc}")
        time.sleep(0.4)
    return saved


def update_workbook(
    rescience_saved: dict[int, str],
    original_saved: dict[int, str],
) -> None:
    """Append a ``pdf_file`` column to each section in the xlsx.

    The xlsx uses a two-row header: row 1 carries the section label (``RESCIENCE C``
    / ``ORIGINAL``) and row 2 carries field names. Data rows start at row 3.
    """
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb.active

    # Insert a new column right after rescience_filename (column R, 1-based index 18).
    # openpyxl shifts every column at idx >= 19 one to the right, which keeps the
    # ORIGINAL section intact and hands us a blank column 19 to fill.
    ws.insert_cols(19)
    ws.cell(row=1, column=19, value="RESCIENCE C")
    ws.cell(row=2, column=19, value="pdf_file")

    # After the shift, ORIGINAL's filename lives at column 28. Append one more.
    original_pdf_col = ws.max_column + 1
    ws.cell(row=1, column=original_pdf_col, value="ORIGINAL")
    ws.cell(row=2, column=original_pdf_col, value="pdf_file")

    for idx, fname in rescience_saved.items():
        ws.cell(row=idx + 3, column=19, value=fname)
    for idx, fname in original_saved.items():
        ws.cell(row=idx + 3, column=original_pdf_col, value=fname)

    wb.save(XLSX_PATH)
    print(f"\nUpdated workbook: {XLSX_PATH}")


def main() -> None:
    df = load_table()
    print(f"Loaded {len(df)} entries from {XLSX_PATH.name}\n")

    print("=== RESCIENCE C ===")
    rescience_saved = download_rescience(df)

    print("\n=== ORIGINAL ===")
    original_saved = download_originals(df)

    print(
        f"\nRESCIENCE C: {len(rescience_saved)}/{len(df)} downloaded"
        f"\nORIGINAL:    {len(original_saved)}/{len(df)} downloaded"
    )

    update_workbook(rescience_saved, original_saved)


if __name__ == "__main__":
    main()
